from functools import wraps
from datetime import date, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, session, abort, send_file, jsonify)
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from ..db import q, q1, run

bp = Blueprint("admin", __name__)


def admin_req(f):
    @wraps(f)
    def d(*a, **kw):
        if session.get("user", {}).get("role") not in ("admin", "super"):
            abort(403)
        return f(*a, **kw)
    return d

# ── helpers ──────────────────────────────────────────────────────────────────
def _dept_summary():
    today = date.today().isoformat()
    return {
        "active_courses":    q1("SELECT COUNT(*) AS c FROM courses")["c"],
        "sessions_today":    q1("SELECT COUNT(*) AS c FROM class_sessions WHERE session_date=?", (today,))["c"],
        "enrolled_students": q1("SELECT COUNT(*) AS c FROM students WHERE consent_given=1")["c"],
        "total_students":    q1("SELECT COUNT(*) AS c FROM students")["c"],
        "total_lecturers":   q1("SELECT COUNT(*) AS c FROM lecturers")["c"],
        "frt_enrolled":      q1("SELECT COUNT(*) AS c FROM students WHERE enrolled_on_terminal=1")["c"],
        "present_today":     q1("""SELECT COUNT(*) AS c FROM attendance_records ar
                                   JOIN class_sessions cs ON cs.id=ar.session_id
                                   WHERE cs.session_date=? AND ar.status='present'""", (today,))["c"],
        "late_today":        q1("""SELECT COUNT(*) AS c FROM attendance_records ar
                                   JOIN class_sessions cs ON cs.id=ar.session_id
                                   WHERE cs.session_date=? AND ar.status='late'""", (today,))["c"],
        "absent_today":      q1("""SELECT COUNT(*) AS c FROM attendance_records ar
                                   JOIN class_sessions cs ON cs.id=ar.session_id
                                   WHERE cs.session_date=? AND ar.status='absent'""", (today,))["c"],
    }

# ── Dashboard ─────────────────────────────────────────────────────────────────
@bp.route("/")
@admin_req
def dashboard():
    today = date.today().isoformat()
    stats = _dept_summary()

    today_sessions = q("""
        SELECT cs.*, c.code, c.title,
               u.full_name AS lecturer_name,
               COUNT(DISTINCT cr.student_id) FILTER(
                 WHERE cs.group_filter IS NULL OR cs.group_filter=(
                   SELECT group_name FROM students WHERE id=cr.student_id)
               ) AS enrolled,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS cnt_present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS cnt_late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS cnt_absent
        FROM class_sessions cs
        JOIN courses c ON c.id=cs.course_id
        LEFT JOIN lecturers l ON l.id=c.lecturer_id
        LEFT JOIN users u ON u.id=l.user_id
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        LEFT JOIN course_registrations cr ON cr.course_id=cs.course_id
        WHERE cs.session_date=?
        GROUP BY cs.id ORDER BY cs.start_time""", (today,))

    # 7-day trend
    trend = []
    for i in range(6, -1, -1):
        d_ = (date.today() - timedelta(days=i)).isoformat()
        row = q1("""SELECT
            SUM(CASE WHEN ar.status IN ('present','late') THEN 1 ELSE 0 END) AS p,
            COUNT(ar.id) AS t
            FROM class_sessions cs
            JOIN attendance_records ar ON ar.session_id=cs.id
            WHERE cs.session_date=?""", (d_,))
        rate = round(100 * row["p"] / row["t"], 1) if row and row["t"] else 0
        trend.append({"date": (date.today()-timedelta(days=i)).strftime("%a %d"), "rate": rate})

    # Absence alerts
    alerts = q("""
        SELECT s.index_number, s.group_name, s.level, u.full_name,
               c.code AS course_code, c.title AS course_title,
               COUNT(*) AS absent_count
        FROM attendance_records ar
        JOIN students s ON s.id=ar.student_id
        JOIN users u ON u.id=s.user_id
        JOIN class_sessions cs ON cs.id=ar.session_id
        JOIN courses c ON c.id=cs.course_id
        WHERE ar.status='absent'
        GROUP BY ar.student_id, c.id
        HAVING absent_count >= 2
        ORDER BY absent_count DESC LIMIT 8""")

    return render_template("admin/dashboard.html", stats=stats,
                           today_sessions=today_sessions, trend=trend, alerts=alerts)

# ── Students ──────────────────────────────────────────────────────────────────
@bp.route("/students")
@admin_req
def students():
    q_   = request.args.get("q", "").strip()
    grp  = request.args.get("group", "")
    lvl  = request.args.get("level", "")
    dept = request.args.get("dept", "")
    sql  = """SELECT s.*, u.full_name, u.email,
                     p.name AS prog_name, p.code AS prog_code,
                     d.name AS dept_name, d.id AS dept_id,
                     f.code AS fac_code
              FROM students s JOIN users u ON u.id=s.user_id
              LEFT JOIN programmes p ON p.id=s.programme_id
              LEFT JOIN departments d ON d.id=p.department_id
              LEFT JOIN faculties f ON f.id=d.faculty_id
              WHERE 1=1 """
    params = []
    if q_:
        sql += " AND (u.full_name LIKE ? OR s.index_number LIKE ? OR s.employee_id LIKE ?)"
        params += [f"%{q_}%", f"%{q_}%", f"%{q_}%"]
    if grp:
        sql += " AND s.group_name=?"
        params.append(grp)
    if lvl:
        sql += " AND s.level=?"
        params.append(int(lvl))
    if dept:
        sql += " AND d.id=?"
        params.append(int(dept))
    sql += " ORDER BY u.full_name LIMIT 300"
    students = q(sql, params)
    groups   = [r["group_name"] for r in q("SELECT DISTINCT group_name FROM students WHERE group_name IS NOT NULL ORDER BY group_name")]
    depts    = q("SELECT id, name FROM departments ORDER BY name")
    return render_template("admin/students.html", students=students, q=q_,
                           groups=groups, sel_group=grp, sel_level=lvl,
                           sel_dept=dept, departments=depts)


@bp.route("/students/add", methods=["GET", "POST"])
@admin_req
def student_add():
    programmes = q("SELECT p.*, d.name AS dept_name FROM programmes p JOIN departments d ON d.id=p.department_id ORDER BY p.name")
    if request.method == "POST":
        from ..db import hash_pw
        email = request.form["email"].strip().lower()
        full_name = request.form["full_name"].strip()
        index_no  = request.form["index_number"].strip()
        prog_id   = int(request.form["programme_id"]) if request.form.get("programme_id") else None
        level     = int(request.form["level"]) if request.form.get("level") else None
        group     = request.form.get("group_name", "").strip()
        gender    = request.form.get("gender", "")
        hall      = request.form.get("hall", "").strip()
        pw        = request.form.get("password", "Student@2026")
        # Check email unique
        if q1("SELECT id FROM users WHERE email=?", (email,)):
            flash("Email already exists.", "danger")
            return render_template("admin/student_form.html", programmes=programmes, action="add", s=request.form)
        uid = run("INSERT INTO users(email,password_hash,full_name,role,is_active) VALUES(?,?,?,'student',1)",
                  (email, hash_pw(pw), full_name))
        run("""INSERT INTO students(user_id,index_number,employee_id,programme_id,level,
                                    group_name,gender,hall,consent_given)
               VALUES(?,?,?,?,?,?,?,?,0)""",
            (uid, index_no, index_no, prog_id, level, group or None, gender, hall))
        flash(f"Student {full_name} registered successfully.", "success")
        return redirect(url_for("admin.students"))
    return render_template("admin/student_form.html", programmes=programmes, action="add", s={})


@bp.route("/students/<int:sid>/edit", methods=["GET", "POST"])
@admin_req
def student_edit(sid):
    student = q1("""SELECT s.*,u.full_name,u.email,p.name AS prog_name
                    FROM students s JOIN users u ON u.id=s.user_id
                    LEFT JOIN programmes p ON p.id=s.programme_id
                    WHERE s.id=?""", (sid,))
    programmes = q("SELECT p.*, d.name AS dept_name FROM programmes p JOIN departments d ON d.id=p.department_id ORDER BY p.name")
    if not student:
        abort(404)
    if request.method == "POST":
        full_name = request.form["full_name"].strip()
        prog_id   = int(request.form["programme_id"]) if request.form.get("programme_id") else None
        level     = int(request.form["level"]) if request.form.get("level") else None
        group     = request.form.get("group_name","").strip() or None
        gender    = request.form.get("gender","")
        hall      = request.form.get("hall","").strip()
        run("UPDATE users SET full_name=? WHERE id=?", (full_name, student["user_id"]))
        run("UPDATE students SET programme_id=?,level=?,group_name=?,gender=?,hall=? WHERE id=?",
            (prog_id, level, group, gender, hall, sid))
        flash("Student details updated.", "success")
        return redirect(url_for("admin.students"))
    return render_template("admin/student_form.html", programmes=programmes, action="edit", s=student)


@bp.route("/students/<int:sid>/enrol", methods=["POST"])
@admin_req
def enrol_student(sid):
    from ..services.hik import HikvisionTerminal
    student = q1("SELECT s.*,u.full_name FROM students s JOIN users u ON u.id=s.user_id WHERE s.id=?", (sid,))
    if not student or not student["consent_given"]:
        flash("Consent not recorded. Cannot enrol.", "warning")
        return redirect(url_for("admin.students"))
    devices = q("SELECT * FROM hik_devices WHERE is_active=1")
    success = 0
    for dev in devices:
        term = HikvisionTerminal(dev["ip_address"], port=dev["port"] or 80,
                                  username=dev["username"] or "admin",
                                  password=dev["password"] or "", simulate=True)
        ok, msg = term.add_person(student["employee_id"], student["full_name"], student["gender"] or "male")
        run("INSERT INTO enrollment_logs(student_id,device_id,action,success,response_text) VALUES(?,?,?,?,?)",
            (sid, dev["id"], "add_person", 1 if ok else 0, msg))
        if ok:
            success += 1
    if success:
        run("UPDATE students SET enrolled_on_terminal=1 WHERE id=?", (sid,))
        flash(f"Student enrolled on {success} terminal(s).", "success")
    else:
        flash("Enrolment failed. Check device connectivity.", "danger")
    return redirect(url_for("admin.students"))

# ── Lecturers ─────────────────────────────────────────────────────────────────
@bp.route("/lecturers")
@admin_req
def lecturers():
    lecs = q("""SELECT l.*, u.full_name, u.email, u.last_login_at,
                       d.name AS dept_name, f.name AS fac_name,
                       COUNT(DISTINCT c.id) AS course_count
                FROM lecturers l JOIN users u ON u.id=l.user_id
                LEFT JOIN departments d ON d.id=l.department_id
                LEFT JOIN faculties f ON f.id=d.faculty_id
                LEFT JOIN courses c ON c.lecturer_id=l.id
                GROUP BY l.id ORDER BY u.full_name""")
    return render_template("admin/lecturers.html", lecturers=lecs)


@bp.route("/lecturers/add", methods=["GET", "POST"])
@admin_req
def lecturer_add():
    departments = q("SELECT d.*, f.name AS fac_name FROM departments d JOIN faculties f ON f.id=d.faculty_id ORDER BY f.name, d.name")
    if request.method == "POST":
        from ..db import hash_pw
        email     = request.form["email"].strip().lower()
        full_name = request.form["full_name"].strip()
        title     = request.form.get("title", "Dr.")
        staff_id  = request.form["staff_id"].strip()
        dept_id   = int(request.form["department_id"]) if request.form.get("department_id") else None
        pw        = request.form.get("password", "Lecturer@2026")
        if q1("SELECT id FROM users WHERE email=?", (email,)):
            flash("Email already exists.", "danger")
            return render_template("admin/lecturer_form.html", departments=departments, l=request.form)
        uid = run("INSERT INTO users(email,password_hash,full_name,role,is_active) VALUES(?,?,?,'lecturer',1)",
                  (email, hash_pw(pw), full_name))
        run("INSERT INTO lecturers(user_id,staff_id,title,department_id) VALUES(?,?,?,?)",
            (uid, staff_id, title, dept_id))
        flash(f"Lecturer {title} {full_name} added.", "success")
        return redirect(url_for("admin.lecturers"))
    return render_template("admin/lecturer_form.html", departments=departments, l={})

# ── Courses ───────────────────────────────────────────────────────────────────
@bp.route("/courses")
@admin_req
def courses():
    courses = q("""SELECT c.*, u.full_name AS lecturer_name, l.title AS lec_title,
                          d.name AS dept_name,
                          COUNT(DISTINCT cr.student_id) AS student_count,
                          COUNT(DISTINCT cs.id) AS session_count
                   FROM courses c
                   LEFT JOIN lecturers l ON l.id=c.lecturer_id
                   LEFT JOIN users u ON u.id=l.user_id
                   LEFT JOIN departments d ON d.id=c.department_id
                   LEFT JOIN course_registrations cr ON cr.course_id=c.id
                   LEFT JOIN class_sessions cs ON cs.course_id=c.id
                   GROUP BY c.id ORDER BY c.code""")
    return render_template("admin/courses.html", courses=courses)


@bp.route("/courses/add", methods=["GET", "POST"])
@admin_req
def course_add():
    lecturers = q("SELECT l.*, u.full_name, l.title FROM lecturers l JOIN users u ON u.id=l.user_id ORDER BY u.full_name")
    departments = q("SELECT * FROM departments ORDER BY name")
    if request.method == "POST":
        run("""INSERT INTO courses(code,title,credit_hours,semester,academic_year,department_id,lecturer_id)
               VALUES(?,?,?,?,?,?,?)""",
            (request.form["code"].strip().upper(),
             request.form["title"].strip(),
             int(request.form.get("credit_hours", 3)),
             int(request.form.get("semester", 2)),
             request.form.get("academic_year", "2025/2026"),
             int(request.form["department_id"]) if request.form.get("department_id") else None,
             int(request.form["lecturer_id"]) if request.form.get("lecturer_id") else None))
        flash("Course added.", "success")
        return redirect(url_for("admin.courses"))
    return render_template("admin/course_form.html", lecturers=lecturers, departments=departments, c={})

# ── Attendance records ────────────────────────────────────────────────────────
@bp.route("/attendance")
@admin_req
def attendance():
    course_id = request.args.get("course_id", "")
    group     = request.args.get("group", "")
    from_date = request.args.get("from_date", (date.today()-timedelta(days=7)).isoformat())
    to_date   = request.args.get("to_date", date.today().isoformat())
    sql = """
        SELECT ar.*, cs.session_date, cs.start_time, cs.end_time, cs.venue, cs.room,
               cs.group_filter, c.code AS course_code, c.title AS course_title,
               u.full_name AS student_name, s.index_number, s.group_name, s.level,
               p.name AS prog_name, p.code AS prog_code
        FROM attendance_records ar
        JOIN class_sessions cs ON cs.id=ar.session_id
        JOIN courses c ON c.id=cs.course_id
        JOIN students s ON s.id=ar.student_id
        JOIN users u ON u.id=s.user_id
        LEFT JOIN programmes p ON p.id=s.programme_id
        WHERE cs.session_date BETWEEN ? AND ?
    """
    params = [from_date, to_date]
    if course_id:
        sql += " AND c.id=?"
        params.append(int(course_id))
    if group:
        sql += " AND s.group_name=?"
        params.append(group)
    sql += " ORDER BY cs.session_date DESC, cs.start_time, u.full_name LIMIT 500"
    records  = q(sql, params)
    courses  = q("SELECT id, code, title FROM courses ORDER BY code")
    groups   = [r["group_name"] for r in q("SELECT DISTINCT group_name FROM students WHERE group_name IS NOT NULL ORDER BY group_name")]
    summary  = {
        "present": sum(1 for r in records if r["status"]=="present"),
        "late":    sum(1 for r in records if r["status"]=="late"),
        "absent":  sum(1 for r in records if r["status"]=="absent"),
        "total":   len(records),
    }
    return render_template("admin/attendance.html", records=records, courses=courses,
                           groups=groups, summary=summary,
                           sel_course=course_id, sel_group=group,
                           from_date=from_date, to_date=to_date)

# ── Reports ───────────────────────────────────────────────────────────────────
@bp.route("/reports")
@admin_req
def reports():
    report_type = request.args.get("type", "daily")
    today = date.today()

    if report_type == "daily":
        target = request.args.get("date", today.isoformat())
        data = _daily_report(target)
    elif report_type == "weekly":
        week_start = (today - timedelta(days=today.weekday())).isoformat()
        week_start = request.args.get("week_start", week_start)
        data = _weekly_report(week_start)
    else:
        month = request.args.get("month", today.strftime("%Y-%m"))
        data = _monthly_report(month)

    courses = q("SELECT id, code, title FROM courses ORDER BY code")
    return render_template("admin/reports.html", courses=courses,
                           report_type=report_type, data=data,
                           today=today.isoformat(),
                           week_start=request.args.get("week_start", (today-timedelta(days=today.weekday())).isoformat()),
                           month=request.args.get("month", today.strftime("%Y-%m")),
                           sel_date=request.args.get("date", today.isoformat()))


def _daily_report(target_date):
    sessions = q("""
        SELECT cs.*, c.code, c.title, u.full_name AS lecturer_name,
               COUNT(DISTINCT cr.student_id) AS enrolled,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS absent
        FROM class_sessions cs
        JOIN courses c ON c.id=cs.course_id
        LEFT JOIN lecturers l ON l.id=c.lecturer_id
        LEFT JOIN users u ON u.id=l.user_id
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        LEFT JOIN course_registrations cr ON cr.course_id=cs.course_id
        WHERE cs.session_date=?
        GROUP BY cs.id ORDER BY cs.start_time""", (target_date,))
    total_p = sum(s["present"] or 0 for s in sessions)
    total_l = sum(s["late"] or 0 for s in sessions)
    total_a = sum(s["absent"] or 0 for s in sessions)
    return {"sessions": sessions, "total_present": total_p, "total_late": total_l,
            "total_absent": total_a, "date": target_date}


def _weekly_report(week_start):
    week_end = (date.fromisoformat(week_start) + timedelta(days=6)).isoformat()
    days = []
    for i in range(7):
        d_ = (date.fromisoformat(week_start) + timedelta(days=i)).isoformat()
        row = q1("""SELECT
            COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS present,
            COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS late,
            COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS absent,
            COUNT(DISTINCT cs.id) AS sessions
            FROM class_sessions cs
            LEFT JOIN attendance_records ar ON ar.session_id=cs.id
            WHERE cs.session_date=?""", (d_,))
        days.append({"date": d_, "day": date.fromisoformat(d_).strftime("%a"),
                     "present": row["present"] or 0, "late": row["late"] or 0,
                     "absent": row["absent"] or 0, "sessions": row["sessions"] or 0})
    return {"days": days, "week_start": week_start, "week_end": week_end,
            "total_present": sum(d["present"] for d in days),
            "total_late":    sum(d["late"] for d in days),
            "total_absent":  sum(d["absent"] for d in days)}


def _monthly_report(month_str):
    courses = q("""
        SELECT c.id, c.code, c.title, u.full_name AS lecturer_name,
               COUNT(DISTINCT cs.id) AS sessions,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS absent
        FROM courses c
        LEFT JOIN lecturers l ON l.id=c.lecturer_id
        LEFT JOIN users u ON u.id=l.user_id
        LEFT JOIN class_sessions cs ON cs.course_id=c.id
          AND strftime('%Y-%m', cs.session_date)=?
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        GROUP BY c.id ORDER BY c.code""", (month_str,))
    return {"courses": courses, "month": month_str,
            "total_present": sum(c["present"] or 0 for c in courses),
            "total_late":    sum(c["late"] or 0 for c in courses),
            "total_absent":  sum(c["absent"] or 0 for c in courses)}


@bp.route("/reports/<int:cid>/export")
@admin_req
def export_report(cid):
    course = q1("""SELECT c.*,u.full_name AS lecturer_name
                   FROM courses c LEFT JOIN lecturers l ON l.id=c.lecturer_id
                   LEFT JOIN users u ON u.id=l.user_id WHERE c.id=?""", (cid,))
    sessions = q("SELECT * FROM class_sessions WHERE course_id=? ORDER BY session_date, start_time", (cid,))
    registrations = q("""SELECT s.index_number, u.full_name, s.group_name, s.level,
                                p.name AS prog_name, p.code AS prog_code
                          FROM course_registrations cr
                          JOIN students s ON s.id=cr.student_id
                          JOIN users u ON u.id=s.user_id
                          LEFT JOIN programmes p ON p.id=s.programme_id
                          WHERE cr.course_id=? ORDER BY s.group_name, u.full_name""", (cid,))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    navy  = PatternFill("solid", fgColor="003366")
    gold  = PatternFill("solid", fgColor="FFB300")
    green = PatternFill("solid", fgColor="D1FAE5")
    amber = PatternFill("solid", fgColor="FEF3C7")
    red_  = PatternFill("solid", fgColor="FEE2E2")
    white_bold = Font(color="FFFFFF", bold=True)

    ws.merge_cells("A1:J1")
    ws["A1"] = "UNIVERSITY OF PROFESSIONAL STUDIES, ACCRA"
    ws["A1"].font = Font(bold=True, size=13, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"ATTENDANCE REPORT — {course['code']}: {course['title']}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:J3")
    ws["A3"] = f"Lecturer: {course['lecturer_name'] or '—'}  |  Semester {course['semester']}  |  {course['academic_year']}"
    ws["A3"].font = Font(italic=True, size=10)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.append([])
    base_cols = ["Index No.", "Student Name", "Programme", "Group", "Level"]
    sess_hdrs = [f"{s['session_date']}\n{s['start_time'][:5]}\n{s['venue'] or ''} {s['room'] or ''}" for s in sessions]
    summ_cols = ["Present", "Late", "Absent", "Rate (%)"]
    ws.append(base_cols + sess_hdrs + summ_cols)
    hrow = ws.max_row
    for cell in ws[hrow]:
        cell.fill = navy
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for reg in registrations:
        row = [reg["index_number"], reg["full_name"],
               reg["prog_code"] or "—", reg["group_name"] or "—",
               f"Level {reg['level']}" if reg["level"] else "—"]
        present = late = absent = 0
        for s in sessions:
            rec = q1("""SELECT status FROM attendance_records
                        WHERE session_id=? AND student_id=(
                          SELECT id FROM students WHERE index_number=?)""",
                     (s["id"], reg["index_number"]))
            st = rec["status"] if rec else "absent"
            row.append(st[0].upper())
            if st == "present":   present += 1
            elif st == "late":    late += 1
            else:                 absent += 1
        total = max(1, len(sessions))
        rate  = round(100 * (present + late) / total, 1)
        row  += [present, late, absent, rate]
        ws.append(row)
        dr = ws.max_row
        for ci, s in enumerate(sessions):
            cell = ws.cell(dr, len(base_cols) + 1 + ci)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = green if cell.value=="P" else (amber if cell.value=="L" else red_)
        ws.cell(dr, len(base_cols)+len(sessions)+4).fill = (
            green if rate >= 75 else (amber if rate >= 60 else red_))

    for col in ws.columns:
        try:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(x.value or "")) for x in col if hasattr(x,"column_letter")), default=10) + 2, 22)
        except Exception:
            pass
    ws.row_dimensions[hrow].height = 45

    out = BytesIO()
    wb.save(out); out.seek(0)
    fname = f"UPSA_{course['code'].replace(' ','_')}_Attendance_{date.today()}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reports/export-summary")
@admin_req
def export_summary():
    """Export daily/weekly/monthly summary as Excel."""
    report_type = request.args.get("type", "daily")
    today = date.today()
    if report_type == "daily":
        target = request.args.get("date", today.isoformat())
        data = _daily_report(target)
        title = f"Daily Report — {target}"
    elif report_type == "weekly":
        ws_ = request.args.get("week_start", (today-timedelta(days=today.weekday())).isoformat())
        data = _weekly_report(ws_)
        title = f"Weekly Report — {data['week_start']} to {data['week_end']}"
    else:
        month = request.args.get("month", today.strftime("%Y-%m"))
        data = _monthly_report(month)
        title = f"Monthly Report — {month}"

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Report"
    navy = PatternFill("solid", fgColor="003366")
    wf   = Font(color="FFFFFF", bold=True)
    ws.merge_cells("A1:G1")
    ws["A1"] = "UNIVERSITY OF PROFESSIONAL STUDIES, ACCRA"
    ws["A1"].font = Font(bold=True, size=13, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")
    ws["A2"] = title; ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True)
    ws.append([])

    if report_type == "weekly":
        ws.append(["Day", "Date", "Sessions", "Present", "Late", "Absent"])
        for cell in ws[ws.max_row]: cell.fill=navy; cell.font=wf
        for d in data["days"]:
            ws.append([d["day"], d["date"], d["sessions"], d["present"], d["late"], d["absent"]])
    elif report_type == "monthly":
        ws.append(["Course Code", "Title", "Sessions", "Present", "Late", "Absent"])
        for cell in ws[ws.max_row]: cell.fill=navy; cell.font=wf
        for c in data["courses"]:
            ws.append([c["code"], c["title"], c["sessions"] or 0, c["present"] or 0, c["late"] or 0, c["absent"] or 0])
    else:
        ws.append(["Course", "Group", "Time", "Enrolled", "Present", "Late", "Absent"])
        for cell in ws[ws.max_row]: cell.fill=navy; cell.font=wf
        for s in data["sessions"]:
            ws.append([s["code"], s["group_filter"] or "All", f"{s['start_time'][:5]}–{s['end_time'][:5]}",
                       s["enrolled"] or 0, s["present"] or 0, s["late"] or 0, s["absent"] or 0])

    for col in ws.columns:
        try:
            ws.column_dimensions[col[0].column_letter].width = max(
                (len(str(x.value or "")) for x in col if hasattr(x,"column_letter")), default=10) + 2
        except Exception:
            pass
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True,
                     download_name=f"UPSA_{report_type}_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Sessions ──────────────────────────────────────────────────────────────────
@bp.route("/sessions")
@admin_req
def sessions():
    filter_date = request.args.get("date", date.today().isoformat())
    sessions = q("""
        SELECT cs.*, c.code, c.title, u.full_name AS lecturer_name,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS cnt_present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS cnt_late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS cnt_absent,
               COUNT(DISTINCT cr.student_id) AS enrolled
        FROM class_sessions cs
        JOIN courses c ON c.id=cs.course_id
        LEFT JOIN lecturers l ON l.id=c.lecturer_id
        LEFT JOIN users u ON u.id=l.user_id
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        LEFT JOIN course_registrations cr ON cr.course_id=cs.course_id
          AND (cs.group_filter IS NULL OR cs.group_filter=(
               SELECT group_name FROM students WHERE id=cr.student_id))
        WHERE cs.session_date=?
        GROUP BY cs.id ORDER BY cs.start_time""", (filter_date,))
    return render_template("admin/sessions.html", sessions=sessions, filter_date=filter_date)

# ── Devices ───────────────────────────────────────────────────────────────────
@bp.route("/devices")
@admin_req
def devices():
    from ..services.hik import HikvisionTerminal
    devs = q("SELECT * FROM hik_devices ORDER BY name")
    statuses = []
    for d in devs:
        term = HikvisionTerminal(d["ip_address"], simulate=True)
        info = term.get_device_info()
        statuses.append(dict(d) | {"info": info, "online": bool(d["is_active"])})
    return render_template("admin/devices.html", statuses=statuses)

# ── Alerts ────────────────────────────────────────────────────────────────────
@bp.route("/alerts")
@admin_req
def alerts():
    alerts = q("""
        SELECT s.index_number, s.group_name, s.level, u.full_name,
               p.name AS prog_name, p.code AS prog_code,
               c.code AS course_code, c.title AS course_title,
               COUNT(*) AS consecutive_absences
        FROM attendance_records ar
        JOIN students s ON s.id=ar.student_id
        JOIN users u ON u.id=s.user_id
        LEFT JOIN programmes p ON p.id=s.programme_id
        JOIN class_sessions cs ON cs.id=ar.session_id
        JOIN courses c ON c.id=cs.course_id
        WHERE ar.status='absent'
        GROUP BY ar.student_id, c.id
        HAVING consecutive_absences >= 2
        ORDER BY consecutive_absences DESC""")
    return render_template("admin/alerts.html", alerts=alerts)
