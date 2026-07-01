from functools import wraps
from datetime import date, timedelta
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, session, abort, jsonify, send_file)
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from ..db import q, q1, run

bp = Blueprint("lecturer", __name__)

def lec_req(f):
    @wraps(f)
    def d(*a, **kw):
        if session.get("user", {}).get("role") != "lecturer":
            abort(403)
        return f(*a, **kw)
    return d

def _me():
    uid = session["user"]["id"]
    return q1("""SELECT l.*, u.full_name, u.email, d.name AS dept_name
                 FROM lecturers l JOIN users u ON u.id=l.user_id
                 LEFT JOIN departments d ON d.id=l.department_id
                 WHERE l.user_id=?""", (uid,))

@bp.route("/")
@lec_req
def dashboard():
    lec = _me()
    if not lec:
        flash("Lecturer profile not found.", "warning")
        return render_template("lecturer/dashboard.html", lec=None,
                               my_courses=[], today_sessions=[], course_students={})
    today = date.today().isoformat()
    my_courses = q("""SELECT c.*, COUNT(DISTINCT cr.student_id) AS student_count,
                             COUNT(DISTINCT cs.id) AS session_count
                      FROM courses c LEFT JOIN course_registrations cr ON cr.course_id=c.id
                      LEFT JOIN class_sessions cs ON cs.course_id=c.id
                      WHERE c.lecturer_id=? GROUP BY c.id ORDER BY c.code""", (lec["id"],))
    today_sessions = q("""
        SELECT cs.*, c.code, c.title, c.id AS course_id,
               COUNT(DISTINCT cr.student_id) FILTER(
                 WHERE cs.group_filter IS NULL OR cs.group_filter=(
                   SELECT group_name FROM students WHERE id=cr.student_id)) AS enrolled,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS cnt_present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS cnt_late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS cnt_absent
        FROM class_sessions cs
        JOIN courses c ON c.id=cs.course_id
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        LEFT JOIN course_registrations cr ON cr.course_id=cs.course_id
        WHERE cs.session_date=? AND c.lecturer_id=?
        GROUP BY cs.id ORDER BY cs.start_time""", (today, lec["id"]))

    # Per-course student class list with overall attendance totals
    course_students = {}
    for course in my_courses:
        students = q("""
            SELECT s.index_number, u.full_name, s.group_name, s.level,
                   p.code AS prog_code, p.name AS prog_name,
                   COUNT(DISTINCT ar.id) FILTER(WHERE ar.status='present') AS present,
                   COUNT(DISTINCT ar.id) FILTER(WHERE ar.status='late')    AS late,
                   COUNT(DISTINCT ar.id) FILTER(WHERE ar.status='absent')  AS absent
            FROM course_registrations cr
            JOIN students s ON s.id=cr.student_id
            JOIN users u ON u.id=s.user_id
            LEFT JOIN programmes p ON p.id=s.programme_id
            LEFT JOIN class_sessions cs ON cs.course_id=cr.course_id
              AND (cs.group_filter IS NULL OR cs.group_filter=s.group_name)
            LEFT JOIN attendance_records ar ON ar.session_id=cs.id AND ar.student_id=s.id
            WHERE cr.course_id=?
            GROUP BY s.id ORDER BY s.group_name, u.full_name""", (course["id"],))
        course_students[course["id"]] = students

    return render_template("lecturer/dashboard.html", lec=lec,
                           my_courses=my_courses, today_sessions=today_sessions,
                           course_students=course_students)


@bp.route("/session/<int:sid>/live")
@lec_req
def live(sid):
    lec = _me()
    s = q1("""SELECT cs.*, c.code, c.title, c.lecturer_id
              FROM class_sessions cs JOIN courses c ON c.id=cs.course_id WHERE cs.id=?""", (sid,))
    if not s or s["lecturer_id"] != lec["id"]:
        abort(403)
    # Get full student list for override dropdown
    students = q("""SELECT s.id, s.index_number, u.full_name, s.group_name
                    FROM course_registrations cr
                    JOIN students s ON s.id=cr.student_id
                    JOIN users u ON u.id=s.user_id
                    WHERE cr.course_id=?
                      AND (? IS NULL OR s.group_name=?)
                    ORDER BY u.full_name""",
                 (s["course_id"], s["group_filter"], s["group_filter"]))
    return render_template("lecturer/live.html", session=s, lec=lec, students=students)


@bp.route("/session/<int:sid>/data")
@lec_req
def live_data(sid):
    lec = _me()
    s = q1("SELECT cs.*,c.lecturer_id FROM class_sessions cs JOIN courses c ON c.id=cs.course_id WHERE cs.id=?", (sid,))
    if not s or s["lecturer_id"] != lec["id"]:
        abort(403)
    rows = q("""
        SELECT s.id AS student_id, s.index_number, u.full_name, s.group_name, s.level,
               p.name AS prog_name, p.code AS prog_code,
               d.name AS dept_name,
               ar.status, ar.check_in_time, ar.minutes_late, ar.similarity_score, ar.source
        FROM course_registrations cr
        JOIN students s ON s.id=cr.student_id
        JOIN users u ON u.id=s.user_id
        LEFT JOIN programmes p ON p.id=s.programme_id
        LEFT JOIN departments d ON d.id=p.department_id
        LEFT JOIN attendance_records ar ON ar.session_id=? AND ar.student_id=s.id
        WHERE cr.course_id=?
          AND (? IS NULL OR s.group_name=?)
        ORDER BY u.full_name""",
        (sid, s["course_id"], s["group_filter"], s["group_filter"]))
    data = []
    for r in rows:
        ci = r["check_in_time"]
        data.append({
            "student_id": r["student_id"],
            "index":      r["index_number"],
            "name":       r["full_name"],
            "group":      r["group_name"] or "—",
            "level":      r["level"] or "—",
            "programme":  r["prog_code"] or "—",
            "department": r["dept_name"] or "—",
            "status":     r["status"] or "absent",
            "check_in":   ci[11:16] if ci else None,
            "mins_late":  r["minutes_late"] or 0,
            "score":      round(r["similarity_score"], 3) if r["similarity_score"] else None,
            "source":     r["source"] or "terminal",
        })
    summary = {
        "present": sum(1 for r in data if r["status"]=="present"),
        "late":    sum(1 for r in data if r["status"]=="late"),
        "absent":  sum(1 for r in data if r["status"]=="absent"),
        "total":   len(data),
        "is_open": bool(s["is_open"]),
    }
    att_pct = round(100*(summary["present"]+summary["late"])/max(1,summary["total"]),1)
    summary["att_pct"] = att_pct
    return jsonify({"rows": data, "summary": summary})


@bp.route("/session/<int:sid>/toggle", methods=["POST"])
@lec_req
def toggle(sid):
    lec = _me()
    s = q1("SELECT cs.*,c.lecturer_id FROM class_sessions cs JOIN courses c ON c.id=cs.course_id WHERE cs.id=?", (sid,))
    if not s or s["lecturer_id"] != lec["id"]:
        abort(403)
    new_state = 0 if s["is_open"] else 1
    run("UPDATE class_sessions SET is_open=? WHERE id=?", (new_state, sid))
    flash(f"Session {'opened' if new_state else 'closed'} for attendance.", "info")
    return redirect(url_for("lecturer.live", sid=sid))


@bp.route("/session/<int:sid>/override", methods=["POST"])
@lec_req
def override(sid):
    lec = _me()
    s = q1("SELECT cs.*,c.lecturer_id FROM class_sessions cs JOIN courses c ON c.id=cs.course_id WHERE cs.id=?", (sid,))
    if not s or s["lecturer_id"] != lec["id"]:
        abort(403)
    student_id = int(request.form["student_id"])
    new_status = request.form["status"]
    note = request.form.get("note", "")
    existing = q1("SELECT id FROM attendance_records WHERE session_id=? AND student_id=?", (sid, student_id))
    if existing:
        run("UPDATE attendance_records SET status=?,source='manual',note=? WHERE id=?",
            (new_status, note, existing["id"]))
    else:
        run("INSERT INTO attendance_records(session_id,student_id,status,source,note) VALUES(?,?,?,'manual',?)",
            (sid, student_id, new_status, note))
    flash("Attendance record updated.", "success")
    return redirect(url_for("lecturer.live", sid=sid))


@bp.route("/session/<int:sid>/export")
@lec_req
def export_session(sid):
    lec = _me()
    s = q1("""SELECT cs.*,c.code,c.title,c.lecturer_id FROM class_sessions cs
              JOIN courses c ON c.id=cs.course_id WHERE cs.id=?""", (sid,))
    if not s or s["lecturer_id"] != lec["id"]:
        abort(403)
    rows = q("""
        SELECT s.index_number, u.full_name, s.group_name, s.level,
               p.name AS prog_name, p.code AS prog_code, d.name AS dept_name,
               ar.status, ar.check_in_time, ar.minutes_late, ar.similarity_score
        FROM course_registrations cr
        JOIN students s ON s.id=cr.student_id
        JOIN users u ON u.id=s.user_id
        LEFT JOIN programmes p ON p.id=s.programme_id
        LEFT JOIN departments d ON d.id=p.department_id
        LEFT JOIN attendance_records ar ON ar.session_id=? AND ar.student_id=s.id
        WHERE cr.course_id=? AND (? IS NULL OR s.group_name=?)
        ORDER BY s.group_name, u.full_name""",
        (sid, s["course_id"], s["group_filter"], s["group_filter"]))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Attendance"
    navy  = PatternFill("solid", fgColor="003366")
    green = PatternFill("solid", fgColor="D1FAE5")
    amber = PatternFill("solid", fgColor="FEF3C7")
    red_  = PatternFill("solid", fgColor="FEE2E2")
    wf    = Font(color="FFFFFF", bold=True)

    ws.merge_cells("A1:J1")
    ws["A1"] = "UNIVERSITY OF PROFESSIONAL STUDIES, ACCRA"
    ws["A1"].font = Font(bold=True, size=13, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:J2")
    ws["A2"] = f"{s['code']}: {s['title']} — Session Attendance Sheet"
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:J3")
    ws["A3"] = f"Date: {s['session_date']}  |  Time: {s['start_time'][:5]}–{s['end_time'][:5]}  |  Venue: {(s['venue'] or '')} {(s['room'] or '')}  |  Group: {s['group_filter'] or 'All'}"
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(italic=True)
    ws.append([])

    headers = ["Index No.", "Full Name", "Programme", "Department", "Group", "Level",
               "Status", "Time In", "Mins Late", "Similarity (%)"]
    ws.append(headers)
    for cell in ws[ws.max_row]: cell.fill=navy; cell.font=wf; cell.alignment=Alignment(horizontal="center")

    for r in rows:
        st = r["status"] or "absent"
        ci = r["check_in_time"][11:16] if r["check_in_time"] else "—"
        sim = f"{r['similarity_score']*100:.0f}" if r["similarity_score"] else "—"
        ws.append([r["index_number"], r["full_name"],
                   r["prog_code"] or "—", r["dept_name"] or "—",
                   r["group_name"] or "—", f"Level {r['level']}" if r["level"] else "—",
                   st.upper(), ci, r["minutes_late"] or 0, sim])
        fill = green if st=="present" else (amber if st=="late" else red_)
        ws.cell(ws.max_row, 7).fill = fill

    for col in ws.columns:
        try:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = max(
                (len(str(cell.value or "")) for cell in col if hasattr(cell,'value')), default=12) + 2
        except Exception:
            pass

    out = BytesIO(); wb.save(out); out.seek(0)
    fname = f"UPSA_{s['code'].replace(' ','_')}_{s['session_date']}_Group{s['group_filter'] or 'All'}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/reports")
@lec_req
def reports():
    lec = _me()
    my_sessions = q("""
        SELECT cs.*, c.code, c.title, c.id AS course_id,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='present') AS cnt_present,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='late')    AS cnt_late,
               COUNT(DISTINCT ar.student_id) FILTER(WHERE ar.status='absent')  AS cnt_absent,
               COUNT(DISTINCT cr.student_id) AS enrolled
        FROM class_sessions cs
        JOIN courses c ON c.id=cs.course_id
        LEFT JOIN attendance_records ar ON ar.session_id=cs.id
        LEFT JOIN course_registrations cr ON cr.course_id=cs.course_id
          AND (cs.group_filter IS NULL OR cs.group_filter=(SELECT group_name FROM students WHERE id=cr.student_id))
        WHERE c.lecturer_id=? AND cs.session_date<=date('now')
        GROUP BY cs.id ORDER BY cs.session_date DESC LIMIT 60""", (lec["id"],))
    my_courses = q("""SELECT c.*, COUNT(DISTINCT cr.student_id) AS student_count,
                             COUNT(DISTINCT cs.id) AS session_count
                      FROM courses c LEFT JOIN course_registrations cr ON cr.course_id=c.id
                      LEFT JOIN class_sessions cs ON cs.course_id=c.id
                      WHERE c.lecturer_id=? GROUP BY c.id ORDER BY c.code""", (lec["id"],))
    return render_template("lecturer/reports.html", sessions=my_sessions, lec=lec, my_courses=my_courses)


@bp.route("/course/<int:cid>/export")
@lec_req
def export_course(cid):
    """Export full course attendance — accessible by the course's lecturer."""
    lec = _me()
    course = q1("""SELECT c.*,u.full_name AS lecturer_name
                   FROM courses c LEFT JOIN lecturers l ON l.id=c.lecturer_id
                   LEFT JOIN users u ON u.id=l.user_id
                   WHERE c.id=? AND c.lecturer_id=?""", (cid, lec["id"]))
    if not course:
        abort(403)

    sessions = q("SELECT * FROM class_sessions WHERE course_id=? ORDER BY session_date, start_time", (cid,))
    registrations = q("""SELECT s.index_number, u.full_name, s.group_name, s.level,
                                p.name AS prog_name, p.code AS prog_code, d.name AS dept_name
                          FROM course_registrations cr
                          JOIN students s ON s.id=cr.student_id
                          JOIN users u ON u.id=s.user_id
                          LEFT JOIN programmes p ON p.id=s.programme_id
                          LEFT JOIN departments d ON d.id=p.department_id
                          WHERE cr.course_id=? ORDER BY s.group_name, u.full_name""", (cid,))

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from io import BytesIO
    from datetime import date
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    navy  = PatternFill("solid", fgColor="002147")
    green = PatternFill("solid", fgColor="D1FAE5")
    amber = PatternFill("solid", fgColor="FEF3C7")
    red_  = PatternFill("solid", fgColor="FFE4E4")
    wf    = Font(color="FFFFFF", bold=True)

    ws.merge_cells("A1:K1")
    ws["A1"] = "UNIVERSITY OF PROFESSIONAL STUDIES, ACCRA"
    ws["A1"].font = Font(bold=True, size=13, color="002147")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:K2")
    ws["A2"] = f"ATTENDANCE REPORT — {course['code']}: {course['title']}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:K3")
    ws["A3"] = f"Lecturer: {course['lecturer_name'] or '—'}  |  Semester {course['semester']}  |  {course['academic_year']}"
    ws["A3"].font = Font(italic=True, size=10)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.append([])

    base_cols = ["Index No.", "Student Name", "Programme", "Department", "Group", "Level"]
    sess_hdrs = [f"{s['session_date']}\n{s['start_time'][:5]}" for s in sessions]
    ws.append(base_cols + sess_hdrs + ["Present", "Late", "Absent", "Rate (%)"])
    hrow = ws.max_row
    for cell in ws[hrow]:
        cell.fill = navy
        cell.font = wf
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for reg in registrations:
        row = [reg["index_number"], reg["full_name"],
               reg["prog_code"] or "—", reg["dept_name"] or "—",
               reg["group_name"] or "—",
               f"Level {reg['level']}" if reg["level"] else "—"]
        present = late = absent = 0
        for s in sessions:
            rec = q1("""SELECT status FROM attendance_records
                        WHERE session_id=? AND student_id=(
                          SELECT id FROM students WHERE index_number=?)""",
                     (s["id"], reg["index_number"]))
            st = rec["status"] if rec else "absent"
            row.append(st[0].upper())
            if st == "present":   present += 1
            elif st == "late":    late += 1
            else:                 absent += 1
        total = max(1, len(sessions))
        rate  = round(100 * (present + late) / total, 1)
        row  += [present, late, absent, rate]
        ws.append(row)
        dr = ws.max_row
        for ci, s in enumerate(sessions):
            cell = ws.cell(dr, len(base_cols) + 1 + ci)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = green if cell.value == "P" else (amber if cell.value == "L" else red_)
        rate_cell = ws.cell(dr, len(base_cols) + len(sessions) + 4)
        rate_cell.fill = green if rate >= 75 else (amber if rate >= 60 else red_)

    for col in ws.columns:
        try:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or "")) for c in col if hasattr(c, "value")), default=10) + 2, 22)
        except Exception:
            pass
    ws.row_dimensions[hrow].height = 40

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"UPSA_{course['code'].replace(' ','_')}_Full_Attendance_{date.today()}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
